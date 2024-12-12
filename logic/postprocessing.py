import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from collections import Counter

def postprocess_results(results):
    """
    Enhance the Excel workbook with a simple summary and a clean percentage-only chart.
    `results` is expected to be a list of lists like:
    [
        ["File_Name", "Subject", "Grade", "Justification", "Student_Feedback"],
        ["file1.pdf", "English", "5", "Good reasoning...", "Try to elaborate more on examples..."],
        ...
    ]

    Changes:
    - No cumulative percentages.
    - Only a percentage column relative to total count.
    - A single bar chart showing percentage per grade.
    - The "Justification" and "Student_Feedback" columns are fixed width and wrap text.
    - Grades are stored and formatted as numbers.
    """

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    # Identify the Grade column index from the header
    header = results[0]
    grade_idx = header.index("Grade")

    # Write the raw data (including header)
    for i, row in enumerate(results):
        if i == 0:
            # First row is the header
            ws.append(row)
        else:
            # For data rows, convert the Grade value to float if possible
            new_row = list(row)
            try:
                new_row[grade_idx] = float(new_row[grade_idx])
            except (ValueError, TypeError):
                pass
            ws.append(new_row)

    # Format header row in "Results"
    header_fill = PatternFill(start_color="D9D9D9", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-size columns in Results (initial sizing)
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 4

    # Adjust the Justification column specifically (if it exists)
    if "Justification" in header:
        justification_index = header.index("Justification") + 1  # +1 because columns start at 1
        justification_col_letter = get_column_letter(justification_index)
        # Set a fixed width for Justification column
        ws.column_dimensions[justification_col_letter].width = 90
        # Wrap text in that column
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=justification_index)
            cell.alignment = Alignment(wrap_text=True)

    # Adjust the Student_Feedback column similarly (if it exists)
    if "Student_Feedback" in header:
        feedback_index = header.index("Student_Feedback") + 1
        feedback_col_letter = get_column_letter(feedback_index)
        # Set a fixed width for Student_Feedback column
        ws.column_dimensions[feedback_col_letter].width = 90
        # Wrap text in that column
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=feedback_index)
            cell.alignment = Alignment(wrap_text=True)

    # Apply a numeric format to the Grade column
    grade_col_letter = get_column_letter(grade_idx + 1)
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=grade_idx + 1)
        # If it's a number, format it
        if isinstance(cell.value, (int, float)):
            cell.number_format = '0'

    # Create a named table for the results data (pivot-ready)
    num_rows = ws.max_row
    num_cols = ws.max_column
    table_ref = f"A1:{get_column_letter(num_cols)}{num_rows}"
    results_table = Table(displayName="GradesTable", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    results_table.tableStyleInfo = style
    ws.add_table(results_table)

    # Create a summary sheet
    summary_ws = wb.create_sheet("Summary")

    # Extract numeric grades
    grade_rows = [row for row in results[1:]
                  if len(row) > grade_idx
                  and row[grade_idx] is not None
                  and str(row[grade_idx]).strip() != ""]
    parsed_grades = []
    for row_data in grade_rows:
        try:
            g = float(row_data[grade_idx])
            parsed_grades.append(g)
        except ValueError:
            # skip non-numeric
            pass

    parsed_grades.sort()
    grade_counts = Counter(parsed_grades)
    total_count = len(parsed_grades)

    # Instructions
    summary_ws.append(["How to Use Interactivity with Pivot Tables"])
    summary_ws.append(["1. Go to the 'Results' sheet."])
    summary_ws.append(["2. Select any cell within 'GradesTable'."])
    summary_ws.append(["3. Insert a PivotTable (Insert > PivotTable in Excel)."])
    summary_ws.append(["4. Drag 'Grade', 'Subject', etc. as needed."])
    summary_ws.append(["5. Use PivotCharts and Slicers for interactive filtering/visualization."])
    summary_ws.append([])
    summary_ws.append(["Below is a static summary and chart for quick reference:"])
    summary_ws.append([])  # Blank line

    start_row = summary_ws.max_row + 1

    # Write summary header
    summary_ws.append(["Grade", "Count", "Percentage"])
    # Style the summary header
    for cell in summary_ws[start_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Compute percentages and write data
    for grade in sorted(grade_counts.keys()):
        count = grade_counts[grade]
        percentage = (count / total_count) if total_count > 0 else 0
        summary_ws.append([grade, count, percentage])

    # Auto-size columns in Summary
    for column_cells in summary_ws.columns:
        length = max((len(str(cell.value)) if cell.value is not None else 0) for cell in column_cells)
        summary_ws.column_dimensions[column_cells[0].column_letter].width = length + 4

    # Format Percentage as percentage
    data_start = start_row + 1
    data_end = summary_ws.max_row
    for row_cells in summary_ws.iter_rows(min_row=data_start, min_col=3, max_col=3):
        for cell in row_cells:
            cell.number_format = '0.00%'

    # Create a percentage bar chart
    # Categories: Grades
    cats = Reference(summary_ws, min_col=1, min_row=data_start, max_row=data_end)
    # Data: Percentage (including header for series name)
    percentage_data = Reference(summary_ws, min_col=3, min_row=start_row, max_row=data_end)

    chart = BarChart()
    chart.title = "Percentage of Total Submissions by Grade"
    chart.x_axis.title = "Grade"
    chart.y_axis.title = "Percentage of Total"
    chart.add_data(percentage_data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = openpyxl.chart.label.DataLabelList()
    chart.dataLabels.showVal = True  # Show percentage values on bars

    # Place the chart
    summary_ws.add_chart(chart, f"E{start_row}")

    summary_ws.append([])
    summary_ws.append(["Note: The 'Results' sheet is pivot-ready. Insert a PivotTable to explore interactively!"])
    note_row = summary_ws.max_row
    summary_ws.cell(row=note_row, column=1).font = Font(italic=True, color="808080")

    return wb
